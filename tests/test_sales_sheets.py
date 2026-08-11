from io import BytesIO

from openpyxl import load_workbook


def _new_sale(client, total="321.45"):
    return client.post("/planilhas-vendas/registros", json={
        "tipo_documento": "NOTA_FISCAL",
        "numero_documento": "NF-GRADE-001",
        "data_venda": "2026-08-02",
        "cliente_nome": "CLIENTE GRADE TESTE",
        "quantidade": "2",
        "descricao_original": "CINTA TUBULAR 3T 2M",
        "descricao_padronizada": "CINTA TUBULAR - 3 T - 2 M",
        "familia": "CINTA TUBULAR",
        "valor_unitario": "160.725",
        "valor_total": total,
        "desconto": "0",
        "percentual_desconto": "0",
    })


def test_planilha_cria_edita_e_atualiza_relatorio(admin_client):
    created = _new_sale(admin_client)
    assert created.status_code == 201, created.text
    record = created.json()
    assert record["criado_manual"] is True
    assert record["status_padronizacao"] == "PADRONIZADO"

    updated = admin_client.patch(
        f"/planilhas-vendas/registros/{record['id']}",
        json={"valor_total": "400.00", "cliente_nome": "CLIENTE GRADE ALTERADO"},
    )
    assert updated.status_code == 200, updated.text
    assert float(updated.json()["valor_total"]) == 400

    listed = admin_client.get(
        "/planilhas-vendas/registros?tipo_documento=NOTA_FISCAL"
        "&busca=CLIENTE%20GRADE%20ALTERADO"
    )
    assert [item["id"] for item in listed.json()] == [record["id"]]

    revenue = admin_client.get(
        "/relatorios/faturamento?ano=2026&mes=8&tipo_documento=NOTA_FISCAL"
    )
    assert float(revenue.json()["nota_fiscal"]) >= 400

    history = admin_client.get(
        f"/planilhas-vendas/historico?registro_id={record['id']}"
    ).json()
    assert {item["acao"] for item in history} >= {"CRIADO", "EDITADO"}


def test_planilha_lixeira_restauracao_e_excel(admin_client):
    record = _new_sale(admin_client, "222.00").json()
    deleted = admin_client.delete(f"/planilhas-vendas/registros/{record['id']}")
    assert deleted.status_code == 204
    active = admin_client.get(
        "/planilhas-vendas/registros?tipo_documento=NOTA_FISCAL"
        "&busca=NF-GRADE-001"
    ).json()
    assert all(item["id"] != record["id"] for item in active)
    trash = admin_client.get(
        "/planilhas-vendas/registros?tipo_documento=NOTA_FISCAL"
        "&busca=NF-GRADE-001&lixeira=true"
    ).json()
    assert record["id"] in [item["id"] for item in trash]

    restored = admin_client.post(
        f"/planilhas-vendas/registros/{record['id']}/restaurar"
    )
    assert restored.status_code == 200

    exported = admin_client.get(
        "/planilhas-vendas/exportar/NOTA_FISCAL.xlsx?ano=2026&mes=8"
    )
    assert exported.status_code == 200
    workbook = load_workbook(BytesIO(exported.content), data_only=False)
    assert workbook.sheetnames == ["GERAL"]
    values = list(workbook["GERAL"].values)
    assert values[0][0:3] == ("DATA", "DOCUMENTO", "CLIENTE")
    assert any(row[1] == "NF-GRADE-001" for row in values[1:])
    sale_row = next(
        index for index, row in enumerate(values[1:], 2)
        if row[1] == "NF-GRADE-001"
    )
    sheet = workbook["GERAL"]
    assert sheet[f"J{sale_row}"].value == (
        f'=IFERROR((K{sale_row}+L{sale_row})/F{sale_row},0)'
    )
    assert sheet[f"M{sale_row}"].value == (
        f'=IFERROR(L{sale_row}/(K{sale_row}+L{sale_row}),0)'
    )
    assert sheet[f"N{sale_row}"].value == f'=K{sale_row}+L{sale_row}'
    assert sheet[f"O{sale_row}"].value == (
        f'=IFERROR(K{sale_row}/F{sale_row},0)'
    )
    assert sheet[f"M{sale_row}"].number_format == "0.00%"
    assert "TabelaVendas" in sheet.tables


def test_restaurar_versao_anterior(admin_client):
    record = _new_sale(admin_client, "111.00").json()
    admin_client.patch(
        f"/planilhas-vendas/registros/{record['id']}",
        json={"valor_total": "999.00"},
    )
    history = admin_client.get(
        f"/planilhas-vendas/historico?registro_id={record['id']}"
    ).json()
    edit = next(item for item in history if item["acao"] == "EDITADO")
    restored = admin_client.post(
        f"/planilhas-vendas/historico/{edit['id']}/restaurar"
    )
    assert restored.status_code == 200
    assert float(restored.json()["valor_total"]) == 111
