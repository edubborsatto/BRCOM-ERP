"""Planilhas de vendas editáveis, com histórico e exportação para Excel."""

import hashlib
import io
import json
from datetime import date, datetime
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import extract, func, or_
from sqlalchemy.orm import Session

from app import models, schemas
from app.database import get_db
from app.dependencies import require_permission
from app.import_service import _row_hash, normalizar_texto


router = APIRouter(prefix="/planilhas-vendas", tags=["Planilhas de vendas"])

EDITABLE_FIELDS = (
    "numero_documento", "data_venda", "cliente_nome", "cliente_codigo",
    "contato", "quantidade", "descricao_original", "descricao_padronizada",
    "familia", "valor_unitario", "valor_total", "desconto",
    "percentual_desconto", "observacoes",
)


def _accepted(db: Session):
    return db.query(models.RegistroVendaImportado).join(
        models.ImportacaoPlanilha
    ).filter(
        models.ImportacaoPlanilha.status == "CONFIRMADA",
        or_(
            models.RegistroVendaImportado.status_importacao == "NOVO",
            models.RegistroVendaImportado.decisao_duplicidade == "IMPORTAR",
        ),
    )


def _json_value(value):
    if isinstance(value, (date, datetime, Decimal)):
        return str(value)
    return value


def _snapshot(record: models.RegistroVendaImportado) -> dict:
    return {
        field: _json_value(getattr(record, field))
        for field in EDITABLE_FIELDS
    } | {"ativo": record.ativo}


def _history(
    db: Session,
    record: models.RegistroVendaImportado,
    action: str,
    user: models.Usuario,
    before: dict | None,
    after: dict | None,
):
    db.add(models.HistoricoPlanilhaVenda(
        registro_id=record.id,
        acao=action,
        dados_anteriores=json.dumps(before, ensure_ascii=False) if before else None,
        dados_novos=json.dumps(after, ensure_ascii=False) if after else None,
        usuario_id=user.id,
        usuario_nome=user.nome,
    ))


def _manual_batch(db: Session, document_type: str, user: models.Usuario):
    filename = "Lançamentos manuais - Notas fiscais" if document_type == "NOTA_FISCAL" else "Lançamentos manuais - Recibos"
    batch = db.query(models.ImportacaoPlanilha).filter_by(
        nome_arquivo=filename,
        tipo_documento=document_type,
        status="CONFIRMADA",
    ).first()
    if batch:
        return batch
    batch = models.ImportacaoPlanilha(
        nome_arquivo=filename,
        tipo_documento=document_type,
        aba_origem="GERAL" if document_type == "NOTA_FISCAL" else "PRINCIPAL",
        hash_arquivo=hashlib.sha256(f"manual:{document_type}".encode()).hexdigest(),
        status="CONFIRMADA",
        usuario_id=user.id,
        usuario_nome=user.nome,
        confirmado_em=datetime.now(),
    )
    db.add(batch)
    db.flush()
    return batch


@router.get("/registros", response_model=list[schemas.RegistroImportadoResponse])
def list_records(
    tipo_documento: str = Query(pattern="^(NOTA_FISCAL|RECIBO)$"),
    busca: str | None = None,
    ano: int | None = None,
    mes: int | None = Query(default=None, ge=1, le=12),
    ordenar_por: str = Query(default="data", pattern="^(data|valor|cliente|produto|documento|atualizacao)$"),
    ordem: str = Query(default="desc", pattern="^(asc|desc)$"),
    lixeira: bool = False,
    limite: int = Query(default=1000, ge=1, le=5000),
    db: Session = Depends(get_db),
    _=Depends(require_permission("pode_editar_planilhas")),
):
    query = _accepted(db).filter(
        models.RegistroVendaImportado.tipo_documento == tipo_documento,
        models.RegistroVendaImportado.ativo.is_(not lixeira),
    )
    if ano:
        query = query.filter(extract("year", models.RegistroVendaImportado.data_venda) == ano)
    if mes:
        query = query.filter(extract("month", models.RegistroVendaImportado.data_venda) == mes)
    if busca:
        like = f"%{busca.strip()}%"
        query = query.filter(or_(
            models.RegistroVendaImportado.cliente_nome.ilike(like),
            models.RegistroVendaImportado.descricao_original.ilike(like),
            models.RegistroVendaImportado.descricao_padronizada.ilike(like),
            models.RegistroVendaImportado.numero_documento.ilike(like),
        ))
    fields = {
        "data": models.RegistroVendaImportado.data_venda,
        "valor": models.RegistroVendaImportado.valor_total,
        "cliente": models.RegistroVendaImportado.cliente_nome,
        "produto": models.RegistroVendaImportado.descricao_padronizada,
        "documento": models.RegistroVendaImportado.numero_documento,
        "atualizacao": models.RegistroVendaImportado.atualizado_em,
    }
    field = fields[ordenar_por]
    ordering = field.asc() if ordem == "asc" else field.desc()
    return query.order_by(ordering, models.RegistroVendaImportado.id.desc()).limit(limite).all()


@router.post("/registros", response_model=schemas.RegistroImportadoResponse, status_code=201)
def create_record(
    data: schemas.PlanilhaVendaCreate,
    db: Session = Depends(get_db),
    user: models.Usuario = Depends(require_permission("pode_editar_planilhas")),
):
    batch = _manual_batch(db, data.tipo_documento, user)
    line = (db.query(func.max(models.RegistroVendaImportado.linha_origem)).filter_by(
        importacao_id=batch.id
    ).scalar() or 1) + 1
    values = data.model_dump()
    description = values["descricao_original"].strip()
    values["descricao_padronizada"] = (
        values.get("descricao_padronizada") or normalizar_texto(description)
    )
    row_hash = _row_hash(values)
    record = models.RegistroVendaImportado(
        importacao_id=batch.id,
        linha_origem=line,
        hash_registro=row_hash,
        hash_duplicidade=row_hash,
        status_importacao="NOVO",
        status_padronizacao="PADRONIZADO",
        criado_manual=True,
        atualizado_por_id=user.id,
        atualizado_por_nome=user.nome,
        **values,
    )
    db.add(record)
    db.flush()
    batch.total_linhas += 1
    batch.linhas_novas += 1
    _history(db, record, "CRIADO", user, None, _snapshot(record))
    db.commit()
    db.refresh(record)
    return record


@router.patch("/registros/{record_id}", response_model=schemas.RegistroImportadoResponse)
def update_record(
    record_id: int,
    data: schemas.PlanilhaVendaUpdate,
    db: Session = Depends(get_db),
    user: models.Usuario = Depends(require_permission("pode_editar_planilhas")),
):
    record = _accepted(db).filter(models.RegistroVendaImportado.id == record_id).first()
    if not record or not record.ativo:
        raise HTTPException(404, "Linha não encontrada")
    changes = data.model_dump(exclude_unset=True)
    if not changes:
        return record
    required = {
        "data_venda", "cliente_nome", "quantidade", "descricao_original",
        "descricao_padronizada", "valor_total",
    }
    if any(field in required and value in (None, "") for field, value in changes.items()):
        raise HTTPException(422, "Esta célula é obrigatória e não pode ficar vazia")
    before = _snapshot(record)
    for field, value in changes.items():
        setattr(record, field, value)
    record.atualizado_em = datetime.now()
    record.atualizado_por_id = user.id
    record.atualizado_por_nome = user.nome
    record.hash_duplicidade = _row_hash({
        "tipo_documento": record.tipo_documento,
        "numero_documento": record.numero_documento,
        "data_venda": record.data_venda,
        "cliente_nome": record.cliente_nome,
        "quantidade": record.quantidade,
        "descricao_original": record.descricao_original,
        "valor_total": record.valor_total,
    })
    _history(db, record, "EDITADO", user, before, _snapshot(record))
    db.commit()
    db.refresh(record)
    return record


@router.delete("/registros/{record_id}", status_code=204)
def delete_record(
    record_id: int,
    db: Session = Depends(get_db),
    user: models.Usuario = Depends(require_permission("pode_editar_planilhas")),
):
    record = _accepted(db).filter(models.RegistroVendaImportado.id == record_id).first()
    if not record or not record.ativo:
        raise HTTPException(404, "Linha não encontrada")
    before = _snapshot(record)
    record.ativo = False
    record.atualizado_em = datetime.now()
    record.atualizado_por_id = user.id
    record.atualizado_por_nome = user.nome
    _history(db, record, "MOVIDO_PARA_LIXEIRA", user, before, _snapshot(record))
    db.commit()


@router.post("/registros/{record_id}/restaurar", response_model=schemas.RegistroImportadoResponse)
def restore_deleted(
    record_id: int,
    db: Session = Depends(get_db),
    user: models.Usuario = Depends(require_permission("pode_editar_planilhas")),
):
    record = _accepted(db).filter(models.RegistroVendaImportado.id == record_id).first()
    if not record or record.ativo:
        raise HTTPException(404, "Linha não encontrada na lixeira")
    before = _snapshot(record)
    record.ativo = True
    record.atualizado_em = datetime.now()
    record.atualizado_por_id = user.id
    record.atualizado_por_nome = user.nome
    _history(db, record, "RESTAURADO", user, before, _snapshot(record))
    db.commit()
    db.refresh(record)
    return record


@router.get("/historico", response_model=list[schemas.HistoricoPlanilhaResponse])
def history(
    registro_id: int | None = None,
    limite: int = Query(default=100, ge=1, le=500),
    db: Session = Depends(get_db),
    _=Depends(require_permission("pode_editar_planilhas")),
):
    query = db.query(models.HistoricoPlanilhaVenda)
    if registro_id:
        query = query.filter_by(registro_id=registro_id)
    return query.order_by(models.HistoricoPlanilhaVenda.criado_em.desc()).limit(limite).all()


@router.post("/historico/{history_id}/restaurar", response_model=schemas.RegistroImportadoResponse)
def restore_version(
    history_id: int,
    db: Session = Depends(get_db),
    user: models.Usuario = Depends(require_permission("pode_editar_planilhas")),
):
    version = db.get(models.HistoricoPlanilhaVenda, history_id)
    if not version or not version.dados_anteriores:
        raise HTTPException(404, "Versão anterior não disponível")
    record = db.get(models.RegistroVendaImportado, version.registro_id)
    if not record:
        raise HTTPException(404, "Linha não encontrada")
    before = _snapshot(record)
    snapshot = json.loads(version.dados_anteriores)
    for field in EDITABLE_FIELDS:
        if field not in snapshot:
            continue
        value = snapshot[field]
        if field == "data_venda" and value:
            value = date.fromisoformat(value)
        elif field in {"quantidade", "valor_unitario", "valor_total", "desconto", "percentual_desconto"} and value is not None:
            value = Decimal(value)
        setattr(record, field, value)
    record.ativo = snapshot.get("ativo", True)
    record.atualizado_em = datetime.now()
    record.atualizado_por_id = user.id
    record.atualizado_por_nome = user.nome
    _history(db, record, "VERSAO_RESTAURADA", user, before, _snapshot(record))
    db.commit()
    db.refresh(record)
    return record


@router.get("/exportar/{tipo_documento}.xlsx")
def export_excel(
    tipo_documento: str,
    ano: int | None = None,
    mes: int | None = Query(default=None, ge=1, le=12),
    db: Session = Depends(get_db),
    _=Depends(require_permission("pode_editar_planilhas")),
):
    if tipo_documento not in {"NOTA_FISCAL", "RECIBO"}:
        raise HTTPException(404, "Planilha não encontrada")
    query = _accepted(db).filter(
        models.RegistroVendaImportado.tipo_documento == tipo_documento,
        models.RegistroVendaImportado.ativo.is_(True),
    )
    if ano:
        query = query.filter(extract("year", models.RegistroVendaImportado.data_venda) == ano)
    if mes:
        query = query.filter(extract("month", models.RegistroVendaImportado.data_venda) == mes)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "GERAL" if tipo_documento == "NOTA_FISCAL" else "PRINCIPAL"
    headers = [
        "DATA", "DOCUMENTO", "CLIENTE", "CÓD. CLIENTE", "CONTATO", "QTD.",
        "PRODUTO ORIGINAL", "PRODUTO PADRONIZADO", "FAMÍLIA", "VALOR UNIT.",
        "VALOR TOTAL", "DESCONTO (R$)", "% DESCONTO", "VALOR TOTAL BRUTO",
        "VALOR UNIT. LÍQUIDO", "OBSERVAÇÕES",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="047857")
    records = query.order_by(
        models.RegistroVendaImportado.data_venda,
        models.RegistroVendaImportado.id,
    ).all()
    for row_number, row in enumerate(records, 2):
        sheet.append([
            row.data_venda, row.numero_documento, row.cliente_nome,
            row.cliente_codigo, row.contato, float(row.quantidade),
            row.descricao_original, row.descricao_padronizada, row.familia,
            f'=IFERROR((K{row_number}+L{row_number})/F{row_number},0)',
            float(row.valor_total), float(row.desconto),
            f'=IFERROR(L{row_number}/(K{row_number}+L{row_number}),0)',
            f'=K{row_number}+L{row_number}',
            f'=IFERROR(K{row_number}/F{row_number},0)',
            row.observacoes,
        ])
    sheet["J1"].comment = Comment(
        "Fórmula automática: (Valor total + Desconto) / Quantidade.",
        "BRCom ERP",
    )
    sheet["M1"].comment = Comment(
        "Fórmula automática: Desconto / Valor total bruto.",
        "BRCom ERP",
    )
    sheet["N1"].comment = Comment(
        "Fórmula automática: Valor total + Desconto.",
        "BRCom ERP",
    )
    sheet["O1"].comment = Comment(
        "Fórmula automática: Valor total / Quantidade.",
        "BRCom ERP",
    )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:1"
    sheet.print_area = sheet.dimensions
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4
    widths = (13, 16, 30, 16, 22, 11, 38, 38, 24, 17, 17, 17, 16, 19, 19, 35)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for column in (10, 11, 12, 14, 15):
        for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
            for item in cell:
                item.number_format = 'R$ #,##0.00'
    for cell in sheet.iter_cols(min_col=1, max_col=1, min_row=2):
        for item in cell:
            item.number_format = "dd/mm/yyyy"
    for cell in sheet.iter_cols(min_col=13, max_col=13, min_row=2):
        for item in cell:
            item.number_format = "0.00%"
    formula_fill = PatternFill("solid", fgColor="DDEBF7")
    formula_font = Font(color="1F4E78")
    for column in (10, 13, 14, 15):
        for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
            for item in cell:
                item.fill = formula_fill
                item.font = formula_font
    if records:
        table = Table(displayName="TabelaVendas", ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output = io.BytesIO()
    workbook.save(output)
    filename = "brcom-notas-fiscais-atualizadas.xlsx" if tipo_documento == "NOTA_FISCAL" else "brcom-recibos-atualizados.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
