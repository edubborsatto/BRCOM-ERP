"""Leitura e padronização das planilhas históricas de notas e recibos."""

from __future__ import annotations

import hashlib
import io
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app import models


SHEETS = {"NOTA_FISCAL": "GERAL", "RECIBO": "PRINCIPAL"}
DOCUMENT_COLUMN_INDEX = 1  # Coluna B nos arquivos oficiais da Brasil Comercial.

ALIASES = {
    "cliente_nome": {"cliente", "customer name", "nome cliente", "razao social"},
    "cliente_codigo": {"customer id", "codigo cliente", "cod cliente", "id cliente"},
    "contato": {"contact person", "contato", "pessoa contato", "responsavel da compra", "responsavel"},
    "numero_documento": {"nota", "num n f", "numero nf", "n f", "nf", "numero nota", "numero recibo", "recibo"},
    "data_venda": {"data", "transaction date", "data venda", "emissao", "data emissao"},
    "quantidade": {"qtd", "qtd.", "quantidade", "quantity", "qtde"},
    "descricao_original": {"produto", "descricao", "descricao produto", "item description", "item", "mercadoria"},
    "valor_unitario": {"valor", "ind", "valor ind", "valor unitario", "unit price", "preco unitario", "vl unit"},
    "valor_unitario_item": {"valor ind c desc", "valor individual c desc", "valor und c desc", "ind c desc"},
    "valor_total": {"valor total", "total", "valor total c desc"},
    "valor_total_item": {"prod c desc", "valor prod c desc", "valor produto", "valor produto c desc", "total price"},
    "desconto": {"desconto por produto", "valor desconto"},
    "percentual_desconto": {"desconto", "de desc", "porc de valor", "percentual desconto", "desconto percentual"},
}

FAMILIAS = (
    ("CINTA DE ELEVAÇÃO", ("SLING", "RING", "BAND", "ELEVA")),
    ("CINTA TUBULAR", ("TUBULAR",)),
    ("CINTA DE AMARRAÇÃO", ("AMARRA", "CATRACA")),
    ("MOVIMENTAÇÃO DE CARGAS", ("REBOQUE", "ARRASTE")),
    ("CORRENTES E ACESSÓRIOS", ("CORRENTE", "MANILHA", "GANCHO", "ANEL")),
)


def normalizar_texto(value) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()


def _header(value) -> str:
    return normalizar_texto(value).lower()


def _decimal(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def _date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    return None


def _column_map(sheet, document_type: str) -> tuple[int, dict[str, int]]:
    best = (0, {})
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 35), values_only=True), 1):
        mapping = {}
        for index, value in enumerate(row):
            key = _header(value)
            for canonical, aliases in ALIASES.items():
                if key in aliases and canonical not in mapping:
                    mapping[canonical] = index
        if len(mapping) > len(best[1]):
            best = (row_number, mapping)
    if "valor_unitario_item" in best[1]:
        best[1]["valor_unitario"] = best[1]["valor_unitario_item"]
    if "valor_total_item" in best[1]:
        best[1]["valor_total"] = best[1]["valor_total_item"]
    # Nos dois arquivos oficiais, a coluna B identifica o documento. A regra
    # explícita evita que variações de cabeçalho ("NOTA", "NUM. N.F." ou
    # "RECIBO") deixem a coluna Documento vazia no ERP.
    best[1]["numero_documento"] = DOCUMENT_COLUMN_INDEX
    if "descricao_original" not in best[1] and document_type == "NOTA_FISCAL" and "cliente_nome" in best[1]:
        # No modelo histórico de NF, a empresa aparece numa linha e os itens
        # seguintes usam a mesma coluna CLIENTE para guardar a descrição.
        best[1]["descricao_original"] = best[1]["cliente_nome"]
    required = {"cliente_nome", "data_venda", "quantidade", "descricao_original"}
    missing = required - set(best[1])
    if missing:
        labels = ", ".join(sorted(missing))
        raise HTTPException(422, f"Não encontrei as colunas obrigatórias: {labels}")
    if "valor_total" not in best[1] and "valor_unitario" not in best[1]:
        raise HTTPException(422, "Não encontrei a coluna de valor total ou valor unitário")
    return best


def _extract_number(text: str, patterns: tuple[str, ...]) -> Decimal | None:
    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            return _decimal(match.group(1))
    return None


def _technical_fields(description: str) -> dict:
    upper = unicodedata.normalize("NFKD", str(description or "").upper())
    upper = "".join(char for char in upper if not unicodedata.combining(char))
    upper = re.sub(r"\s+", " ", upper)
    family = None
    for label, words in FAMILIAS:
        if any(word in upper for word in words):
            family = label
            break
    width = _extract_number(upper, (r"(\d+(?:[.,]\d+)?)\s*MM",))
    length = _extract_number(upper, (r"(\d+(?:[.,]\d+)?)\s*(?:MT|MTS|METRO|METROS)\b",))
    capacity = _extract_number(upper, (r"(\d+(?:[.,]\d+)?)\s*(?:TON|TONS|T)\b", r"(\d+(?:[.,]\d+)?)\s*KG\b"))
    hooks = [name for name in ("GANCHO J", "GANCHO D", "GARRA", "MANILHA", "OLHAL") if name in upper]
    return {
        "familia": family,
        "aplicacao": "ELEVAÇÃO" if "ELEVA" in upper or "SLING" in upper else ("AMARRAÇÃO" if "AMARRA" in upper else None),
        "material": "ALLOY" if "ALLOY" in upper else ("POLIÉSTER" if "POLIESTER" in upper else None),
        "largura": width,
        "capacidade": capacity,
        "comprimento": length,
        "gancho": ", ".join(hooks) or None,
    }


def _product_match(db: Session, description: str) -> tuple[models.Produto | None, float]:
    target = normalizar_texto(description)
    target_tokens = set(target.split())
    best_product, best_score = None, 0.0
    for product in db.query(models.Produto).all():
        candidate = normalizar_texto(product.nome)
        tokens = set(candidate.split())
        overlap = len(target_tokens & tokens) / max(1, len(tokens))
        sequence = SequenceMatcher(None, target, candidate).ratio()
        score = max(sequence, overlap * 0.9)
        if candidate == target:
            score = 1.0
        if score > best_score:
            best_product, best_score = product, score
    return best_product, best_score


def _row_hash(data: dict) -> str:
    fields = (
        data["tipo_documento"], data.get("numero_documento"), data["data_venda"],
        normalizar_texto(data["cliente_nome"]), data["quantidade"],
        normalizar_texto(data["descricao_original"]), data["valor_total"],
    )
    return hashlib.sha256("|".join(map(str, fields)).encode()).hexdigest()


def parse_workbook(content: bytes, document_type: str, db: Session) -> tuple[str, list[dict]]:
    expected_sheet = SHEETS.get(document_type)
    if not expected_sheet:
        raise HTTPException(422, "Tipo de documento inválido")
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(422, "O arquivo não é uma planilha Excel .xlsx válida") from exc
    sheet_lookup = {normalizar_texto(name): name for name in workbook.sheetnames}
    real_sheet = sheet_lookup.get(normalizar_texto(expected_sheet))
    if not real_sheet:
        raise HTTPException(422, f'O arquivo deve possuir a aba "{expected_sheet}"')
    sheet = workbook[real_sheet]
    header_row, columns = _column_map(sheet, document_type)
    learned = {}
    previous_records = db.query(models.RegistroVendaImportado).join(models.ImportacaoPlanilha).filter(
        models.ImportacaoPlanilha.status == "CONFIRMADA",
        models.RegistroVendaImportado.ativo.is_(True),
        models.RegistroVendaImportado.status_padronizacao == "PADRONIZADO",
    ).order_by(models.RegistroVendaImportado.id).all()
    for previous in previous_records:
        learned[normalizar_texto(previous.descricao_original)] = previous
    rows, inherited = [], {}
    inherited_fields = ("cliente_nome", "cliente_codigo", "contato", "numero_documento", "data_venda")
    for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        values = {field: row[index] if index < len(row) else None for field, index in columns.items()}
        shared_description = columns["descricao_original"] == columns["cliente_nome"]
        if shared_description:
            shared_value = values.get("cliente_nome")
            if _decimal(values.get("quantidade")) <= 0:
                for field in inherited_fields:
                    if values.get(field) not in (None, ""):
                        inherited[field] = values[field]
                continue
            values["descricao_original"] = shared_value
            values["cliente_nome"] = inherited.get("cliente_nome")
        for field in inherited_fields:
            if values.get(field) not in (None, ""):
                inherited[field] = values[field]
            elif field in inherited:
                values[field] = inherited[field]
        description = str(values.get("descricao_original") or "").strip()
        if not description or normalizar_texto(description) in {"TOTAL", "SUBTOTAL"}:
            continue
        sale_date = _date(values.get("data_venda"))
        customer = str(values.get("cliente_nome") or "").strip()
        quantity = _decimal(values.get("quantidade"))
        if not sale_date or not customer or quantity <= 0:
            continue
        unit_value = _decimal(values.get("valor_unitario"))
        total_value = _decimal(values.get("valor_total")) or unit_value * quantity
        learned_record = learned.get(normalizar_texto(description))
        product, score = _product_match(db, description)
        technical = _technical_fields(description)
        if learned_record:
            standardized = learned_record.descricao_padronizada
        else:
            standardized = product.nome if product and score >= 0.82 else normalizar_texto(description)
        data = {
            "linha_origem": row_number,
            "tipo_documento": document_type,
            "numero_documento": str(values.get("numero_documento") or "").strip() or None,
            "data_venda": sale_date,
            "cliente_nome": customer,
            "cliente_codigo": str(values.get("cliente_codigo") or "").strip() or None,
            "contato": str(values.get("contato") or "").strip() or None,
            "quantidade": quantity,
            "descricao_original": description,
            "descricao_padronizada": standardized,
            "produto_id": learned_record.produto_id if learned_record else (product.id if product and score >= 0.82 else None),
            "valor_unitario": unit_value,
            "valor_total": total_value,
            "desconto": _decimal(values.get("desconto")),
            "percentual_desconto": _decimal(values.get("percentual_desconto")),
            # Uma descrição nova não é um erro: ela entra automaticamente como
            # revisada. A intervenção do gestor fica reservada às duplicidades.
            "status_padronizacao": "PADRONIZADO",
            **technical,
        }
        if learned_record:
            for field in ("familia", "aplicacao", "material", "largura", "capacidade", "comprimento", "gancho", "reforco", "costura", "impressao"):
                data[field] = getattr(learned_record, field) or data.get(field)
        if product and score >= 0.82:
            data["familia"] = product.familia or data["familia"]
            data["largura"] = product.largura or data["largura"]
            data["capacidade"] = product.resistencia or data["capacidade"]
            data["comprimento"] = product.comprimento or data["comprimento"]
        data["hash_registro"] = _row_hash(data)
        rows.append(data)
    if not rows:
        raise HTTPException(422, "Nenhuma linha de venda válida foi encontrada na aba esperada")
    return real_sheet, rows
